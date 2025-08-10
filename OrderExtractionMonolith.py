#NOTES: DO NOT HAVE ANY EXCEL FILES OPEN WHEN STARTING THE SCRIPT - EXCEL DOESN'T SUPPORT CONCURRENT READ-WRITES SO IT WILL THROW AN ERROR
#DO NOT MOVE FILES DURING RUNTIME
#HAVE EVERYTHING INSTALLED
#USE AN EFFICIENT SLM (small language model)

from PIL import Image as Im
import pytesseract
from pytesseract import image_to_data
import io

from pathlib import Path
from pdf2image import convert_from_path
import copy

from img2table.document import Image
from img2table.document import PDF
from img2table.ocr import TesseractOCR
import fitz
import re

import pandas as pd
import openpyxl

import subprocess
import ollama
from ollama import chat
from ollama import ChatResponse
#https://ollama.com/blog/structured-outputs
from pydantic import BaseModel, create_model
import json
from typing import get_type_hints
from enum import Enum

import os
import shutil
import time
import tkinter as tk
from tkinter import filedialog

#If you have any enumerated types (where output MUST be constrained to a few specific choices - think multiple choice or limited options)
#declare them as such. This will let you provide them to the Ollama model for a JSON output
class Attendance(str, Enum):
    Yes = "Yes"
    No = "No"
    WS = "Written submission received"
    WS0 = "Written hearing but no written submission received"

#This should mirror what output you want from the JSON model.
#The types of data are essentially anything Python allows, and enumerated types.
# | acts as OR
class Order(BaseModel):
    Address: str
    FileNumber: str
    LandlordName: str
    DateOfHearing: str
    City: str
    PostalCode: str
    FirstEffectiveDate: str
    WeightedUsefulLife: list[float]
    LandlordRate: float
    AGIRateTotal: float
    CapitalExpenditure: list[str] | None
    LandlordAttendance: Attendance
    LandlordLegalAttendance: Attendance
    TenantAttendance: Attendance
    TenantLegalAttendance: Attendance
    InterestingNotes: str | None

#If you want to preface any messages to the API, put the start in content.
#This is not designated as a true system message though - it simply provides more context if you choose to provide it.
#The role allows the SLM to read what it's provided, from who.
#This will play a larger part if you want conversations.
system_message = {
    "role": 'system',
    "content": "You extract rental increase data from strings, with no extra text."}

#The column names in tables to be extracted from PDFs. These may appear multiple times - we search for them rather than direct matching.
#We normalize them by removing all whitespace
splitColumns = ["TotalIncreaseexcludesguideline",
                "%incforCapExp"]

#The column names in tables to be extracted from PDFs. These should only appear once - we attempt to direct match.
#We normalize them by removing all whitespace
totalColumns = ["WeightedUsefulLifeforCapitalExp*",
                 "Total%forCapExp"]

#These are the SLMs (small language models) you have pulled from Ollama with "ollama pull {model_name}"
#Experiment with what you can run - larger and more recent models are generally better.
models = ["phi3:mini", "qwen3:8b", "deepseek-r1:7b", "llama3.2", "deepseek-r1:1.5b", "qwen3:1.7b"]

#These are parameters that directly affect the tokens chosen by the SLM
#Temperature is essentially a measure of randomness (how much we sample across the most likely next token). Set to 0.0 for near deterministic output.
#Top_p is how many words it should consider for the next token. It is similar to temperature - it controls randomness.
#Repeat penalty is a measure of how strongly we should prefer previous words - 1.0 means no preference
# < 1.0 is to be biased against previous words, > 1.0 is to biased for previous words
#If you KNOW the words/information are in the text, set this high. 5.0 is good.
#num predict is how many tokens long the output should be constrained to. This should be played around with carefully
#If it's too short - the JSON will be incomplete and it throws an error (try telling it that it has a word limit).
#If too high, prepare to let it run for hours with nothing to show.
hyperparameters = {'temperature': 0.0,
                   'top_p': 0.0,
                   'repeat_penalty': 1.0,
                   'num_predict': 4096}

#This is where we turn pdf columns into OUR column names. It's just a simple mapping, but make sure the names are consistent across the program.
possibleColumns = {"WeightedUsefulLifeforCapitalExp*": "WeightedUsefulLife",
                   "Total%forCapExp": "AGIRateTotal",
                   "TotalIncreaseexcludesguideline1": "AGIRate1",
                   "%incforCapExp1": "AGIRate1",
                   "TotalIncreaseexcludesguideline2": "AGIRate2",
                   "%incforCapExp2": "AGIRate2",
                   "TotalIncreaseexcludesguideline3": "AGIRate3",
                   "%incforCapExp3": "AGIRate3",
                   "TotalIncreaseexcludesguideline4": "AGIRate4",
                   "%incforCapExp4": "AGIRate4",
                   "TotalIncreaseexcludesguideline5": "AGIRate5",
                   "%incforCapExp5": "AGIRate5"}

#This is how we do slightly more advanced matches with regular expressions.
#This cannot be easily explained over text - I recommend looking up a guide.
regexDict = {r'(?<=In the matter of:\s)(.*?)(?=\s*Between:)': "Address",
             r'Between:\s*(.*?)\s*Landlord\s*(.*?)\s*and\s*Refer': "LandlordName",
             r'File Number:\s*([A-Z]{3}-\d{5}-\d{2})' : "FileNumber",
             #r'\b([A-Z]\d[A-Z]\d[A-Z]\d)\b(?=[\s\n]*Between:)': "PostalCode",
             r'([A-Z]\d[A-Z]\d[A-Z]\d)': "PostalCode",
             r'This application was heard in (.*?) on': "City",
             r'This application was heard in .+? on (.+?)\.': "DateOfHearing"
             }

otherDict = {r'First Effective Date of Rent Increase in this Order is\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4})': "FirstEffectiveDate",
             r'\b\d+(?:\.\d+)?%': "LandlordRate"}
#in the case anyone needs them and doesn't want to comb through the mess below

#If you want more effective regex or smaller prompts to the SLM,
#this is a way to isolate consistently repeated phrases and their nearby words to pass to another method.
parsingDict = {"LandlordAttendance": "The following parties",
               "LandlordLegalAttendance": "The following parties",
               "TenantAttendance": "The following parties",
               "TenantLegalAttendance": "The following parties",
               "FirstEffectiveDate": "effective date",
               "LandlordRate": "Landlord justified",
               "InterestingNotes": "withdraw",
               "CapitalExpenditure": "apital expenditure"
}

leftoverRegex = {"LandlordAttendance": r"\(The following parties.*?\)(?=\s*It is determined that:)",
                 "LandlordLegalAttendance": r"\(The following parties.*?\)(?=\s*It is determined that:)",
                 "TenantAttendance" : r"\(The following parties.*?\)(?=\s*It is determined that:)",
                 "TenantLegalAttendance": r"\(The following parties.*?\)(?=\s*It is determined that:)"
                 }

#Where we specify (partly for ourselves) to see what method is primarily meant for each field.
fieldsRegex = {"Address", "FileNumber", "LandlordName", "DateOfHearing", "City", "PostalCode"} #regex
miscellaneousFields = {"FirstEffectiveDate", "LandlordRate", "FileName"}
tableFields = {"AGIRateTotal", "AGIRate1", "AGIRate2", "AGIRate3", "AGIRate4", "AGIRate5", "WeightedUsefulLife"} #tables
ollamaFields = {"CapitalExpenditure", "InterestingNotes",
                "LandlordAttendance", "LandlordLegalAttendance","TenantLegalAttendance", "TenantAttendance"} #ollama

#if your fields are ALL here, you can uncomment the following line. You may also make other methods in the future, and simple use the union operation again.
#totalFields = fieldsRegex.union(ollamaFields).union(tableFields)
totalFields = ["FileName", "Address", "LandlordName", "FileNumber", "DateOfHearing", "City", "PostalCode", "FirstEffectiveDate", "LandlordRate",
               "AGIRateTotal", "AGIRate1", "AGIRate2", "AGIRate3", "AGIRate4", "AGIRate5", "WeightedUsefulLife", "CapitalExpenditure", "InterestingNotes",
                "LandlordAttendance", "LandlordLegalAttendance","TenantLegalAttendance", "TenantAttendance"]

#Give the SLM instructions here.
#Be concise, give it any extremely important instructions
#EXPLAIN WHAT EACH FIELD MEANS - it is a token predictor, not something that can read between the lines
questionsContext = "You are an assistant that extracts data from long strings of rent increase court orders." +\
"You keep answers as short as possible (a couple words per field maximum, (one word if possible, your output tokens are limited to 1024) and don't repeat entries for the same field." +\
"The address consists of a number and the street name. It typically ends in 'Road' or 'Street' or 'Drive'.\n" +\
                   "The city is the city of the address which the hearing is concerning." +\
                       "The file number is in the format XXX-DDDDD-DD, where X is a letter and D is a digit." +\
                       "The landlord name immediately precedes '('the Landlord')', and is fully capitalized in the text." +\
                       "The date of hearing is when the hearing was heard. It should be in YYYY-MM-DD format. The city is usually listed with it, but it is not always capitalized." +\
                       "The postal code is usually listed in the address, and it is 6 characters long, in the format XDXDXD or XDX DXD" +\
                       "The first effective date should be in YYYY-MM-DD format, and it is when the rent increase goes into effect."+\
                       "The weighted useful life of capital expenditures should be listed. If there are multiple for different repairs or expenses, include all of them." +\
                       "The landlord rate is the rent increase initially justified by the Landlord." +\
                       "The AGIRateTotal is the total percent increase in rent over all years specified in the text. If it doesn't mention different years, this quantity is just the increase" +\
                       "The AGIRateByYear is an ordered list of the maximum rent increases each year, in percents. If there are n years specified, this should be n elements long." +\
                       "Capital expenditures are the repairs or improvements to the property. If not explicitly stated, use 'Not stated'. If not at all mentioned, use None." +\
                       "The capital expenditures strings should be a couple words maximum." +\
                       "The attendance fields are as such: Yes = 1; No =0; WS = written submission received; WS0 = written hearing but no written submission received. Keep in mind that names may have been censored, so if Tenants or Units are mentioned, that is very likely them being in attendance."+\
                        "If you detect any legal representation, assume that's a 1 or Yes." +\
                       "For fields with legal attendance, assume ANY representation for that party is legal representation." +\
                       "The interesting notes is a list of the following: Were units/tenants removed from the application? Anything else of note? Any pro forma language?"

#Final Approach is what we pass files to and we receive a JSON
def FinalApproach(file):
    #setup - we simple isolate each schedule here. In the future, files may not have schedules - repurpose it there.
    pages = getPagesBySchedule(file)
    #Gets text from schedule 0
    fileContents = getOCR(pages, 0, file)
    #print(fileContents)
    BaseAnswers = {}
    #File name
    BaseAnswers["FileName"] = file.stem
    
    #Regex
    BaseAnswers = Regex(fileContents, BaseAnswers)

    #If you have only one address per file, remove this loop condition
    x = getListOfStringSeparatedThings(BaseAnswers["Address"])
    ConstructedAnswers = []
    firstPass = True
    for address in x:
        answers = copy.deepcopy(BaseAnswers)
        answers["Address"] = address
        answers = Table(file, pages, answers, len(x)!=1)
        
        answers = Miscellaneous(pages, file, fileContents, answers, [FirstEffectiveDate, LandlordRate])

        print("Starting LLMCall")
        answers = LLMOneshot(fileContents, answers)
        print("Finished Oneshot")

        #Very important!!! openpyxl and excel treat some fields differently
        #It tries to 'stringify' each field, but it cannot do these for fields where the underlying C code for Python represents it as non-existent
        #It throws an error, so we simply post-process answers here for an excel-readable JSON.
        #Feel free to do anything you want in this if statement - if it gets too long, consider a case
        for key, value in answers.items():
            if value is None or value == []:
                answers[key] = "FLAGGED FOR REVIEW"
            elif isinstance(value, list):
                answers[key] = ', '.join(map(str, value))  # Turn list into readable string
            elif isinstance(value, dict):
                answers[key] = str(value)
            else:
                answers[key] = value

        ConstructedAnswers.append(answers)
        displayDict(answers)
        
    return ConstructedAnswers

#We try to get all the possible addresses here so we can form every row separately. I expect this to become a relic of this specific PDF type.
#Please consider changing this method or removing it entirely for your use-case
def getListOfStringSeparatedThings(addressString):
    address_pattern = re.compile(
        r'(\d+\s+[^,]+,\s+[^,]+,\s+[A-Z]{2},\s+[A-Z0-9]{6})', re.IGNORECASE)
    
    # Find all complete addresses
    matches = list(address_pattern.finditer(addressString))

    # If nothing found, just pass it back without trailing whitespace
    if not matches:
        return [addressString.strip()] if addressString.strip() else []
    
    result = []

    #Add all regex matches to our address list
    for match in matches:
        address = match.group(1).strip()
        if address:
            result.append(address)
    
    print(result)
    return result

#Again, this is for this specific kind of PDF, and it simply returns the pages that the schedules are on.
def getPagesBySchedule(file):
    pageArray = {0: [], 3: [], 4: [], 1:[], 2:[]}
    fileImages = convert_from_path(file, dpi=400) #dpi is dots per inch, or how we're vectorizing the pdf into an image
    last = 0;

    i = 0
    for image in fileImages: #iterate through the pages
        
        #top strip (from top to -, full width, maybe 200px tall)
        width, height = image.size #.size is how we get the image size
        top_strip = image.crop((0, 0, width, 500)) #we extract a rectangle from pixel coordinate (0,0) to (width, 500)
        #width means spanning the horizontal distance of the pdf, so imagine cutting off the top 1.25 inches of the page and storing it as top_strip
        
        text = pytesseract.image_to_string(top_strip) #use OCR
        text = text.strip()  #cleanup whitespace/newlines


        for j in range(1, 5): #crude loop to check schedule no.
            if "Schedule " + str(last + j) in text:
                last = last + j

        if last in (0,1,2,3,4):
            pageArray[last].append(i) #add it if it's a valid schedule page
        i = i + 1

    print(pageArray)
    return pageArray

""" #Suppose you have a file that doesn't have schedules. Then, simply treat the entire thing as Schedule 0, and change the other methods to operate on Schedule 0.
#Maybe you want to judge schedule by something else, then use similar logic from the above function and build a page classifier
def getPagesBySchedule(file):
    pageArray = {0: []}
    for image in fileImages:
        pageArray[0].append(i)
        i = i + 1
"""
    

#We don't want to pass EVERY field to the model, because then we would overwrite/be longer/less efficient.
#If using the cloud, this also saves on costs.
#We take the model defined (the defined class) and fields to exclude, and build a new class without the excluded fields
def create_filtered_model(base_model, exclude_fields):
    annotations = get_type_hints(base_model)
    
    filtered_annotations = {
        field_name: field_type 
        for field_name, field_type in annotations.items() 
        if field_name not in exclude_fields
    }

    #Don't touch this function. It is syntactically correct.
    return create_model(
        f"{base_model.__name__}Filtered",
        **filtered_annotations
    )

#Given text and fields that we DON'T want, give us the fields that we do want.
#if this is weird, "The missile knows where it is because it knows where it isn't."
def LLMJSON(text, excludeFields):
    #We create the new model from the above method
    FilteredOrder = create_filtered_model(Order, excludeFields)
    #We reboot the model to prevent any possible memory leaks - this is a non-issue currently. 
    subprocess.run(["ollama", "stop", singleModel])

    #We obtain a response using the chat API from ollama. You may choose to add additional messages (it's an array!!! position matters) by adding dictionaries to the message array
    response = chat(
        messages = [
            {'role': 'user',
             'content': text + 
             "Return a JSON object. Be concise and add no extra words.",
            }],
        #Choose the SLM here
        model = singleModel,
        format=FilteredOrder.model_json_schema(),  # Use filtered model schema
        options = hyperparameters, #Provide hyperparameters
        stream=False #Means we ONLY want the response once it's done generating
        )
    return response['message']['content'] #I recommend looking up what the actual return of the chat API is to understand this.

#As the name suggests, we give the SLM one shot to generate the answers
#THIS METHOD DOES A LOT OF THE HEAVY LIFTING IN THIS PROGRAM
def LLMOneshot(text, answers):
    fields_to_exclude = set(answers.keys()) #get fields where we don't have an answer
    try:
        json_response = LLMJSON(text, fields_to_exclude, answers["Address"]) #get JSON
    except Exception: #if any error (output too long, model not installed, etc.) we just use the answers we currently have
        return answers #better than dumping all progress
    parsed_response = json.loads(json_response) #I lied! LLMJSON returns a string that just follows a JSON schema. We turn it into a real JSON dictionary here.

    for key, value in parsed_response.items(): #Populate answer dictionary with our new answers
        answers[key] = value
    return answers

#If you have fields that fall in a weird grey area where they aren't exactly extractable with the other methods
#AND it'd be computationally inefficient to retrieve them with the SLM
#AND you have a clever solution, define them as a method and pass that method in an array to this method
#('clever' is left up to interpretation - someone else might see it as abusing the file structure or other assumptions)
#That method MUST take all file data, but you don't necessarily need to use it all!
#It should update answers within the function scope, but this can be changed if you'd like for better error-handling.
def Miscellaneous(pages, file, fileContents, answers, methods):
    for method in methods: #iterating over the methods
        answers = method(pages, answers, file, fileContents)
    return answers
#My thoughts when I made this - But why would we pass methods??? That makes NO sense!
#Now I think - what if answers extracted are order-dependent? That is, maybe a field is easily extractable GIVEN we have an answer from something else.
#And another field is dependent on THAT output. So on and so forth. This modular approach allows us to easily permute and order our miscellaneous functions.

#This is a miscellaneous function - albeit one that doesn't require previous answers
def FirstEffectiveDate(pages, answers, file, fileContents):
    context = getOCR(pages, 3, file)
    #We just look for the first date that shows up in Schedule 3 (this works because Schedule 3 MUST start with the Ordered Rent Increase date)
    x = regexSmall(context, r'First Effective Date of Rent Increase in this Order is\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4})')
    if x == None:
        x = regexSmall(context, r'((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4})')#just match for ANY date     
    answers["FirstEffectiveDate"] = x
    return answers

#We want the AGI total increase if it's not stated, sum the maximums of the years
def getMax(answers):
    m = 0
    for i in range(1, 6):
        x = (answers["AGIRate"+str(i)]).keys()
        if x:
            m = m + max(x)
    return m

#A 'solution'. Sometimes, orders don't include the rate asked for by the landlord
#How do we get around this? Assuming landlords want as much money as possible! We greedily search for any percentages in Schedule 0 and take the maximum of that and the getMax output
def LandlordRate(pages, answers, file, fileContents):
    context = getOCR(pages, 0, file)
    matches = re.findall(r'\b\d+(?:\.\d+)?%', context)
    numbers = [float(m.strip('%')) for m in matches]
    numbers.append(getMax(answers))
    if numbers:
        answers["LandlordRate"] = max(numbers)
    else:
        answers["LandlordRate"] = 0 #if we have no numbers, we must make the conclusion that this isn't an AGI order at all/contains no reasonable info
    return answers

def AddAGIFields(answers):
    for i in range(5):
        if "AGIRate" + str(i+1) not in answers.keys():
            answers["AGIRate" + str(i+1)] = {} #If we don't have an answer, give it an empty dictionary so we don't get errors.
    return answers

#A wrapper for the Table functions
def Table(file, pageNo, answers, multiple = True): #https://github.com/xavctn/img2table, pretty much exactly what I did and packaged so I'll use this
    if not pageNo[3]:
        print("Warning: No pages found for schedule 3")
        return AddAGIFields(answers)
        
    # Convert Path to string and validate file exists, just in case someone moved it
    file_path = str(file)
    if not Path(file_path).exists():
        raise FileNotFoundError(f"File does not exist: {file_path}")

    start_page = pageNo[3][0]
    end_page = pageNo[3][-1]
    
    pages = list(range(start_page, end_page + 1))

    pdf = PDF(file_path, pages=pages, detect_rotation=False, pdf_text_extraction=True) #turn our pdf into an image
    ocr = TesseractOCR(lang="eng") #specify our ocr engine
    tables = pdf.extract_tables(ocr=ocr) #extract all tables, where it does this by searching for lines and intersections to form cells
    pdf.to_xlsx('tables.xlsx', ocr=ocr)#for testing, throw them into an excel so we can see them
    table = tableJoiner(tables, pages)#pass our list of tables to form a single dataframe - much better for operations
    table.to_excel("output_table.xlsx", index=False)#throw that single dataframe into excel just so we can check

    if multiple:
        answers = tableToFields(table, answers)
    else:
        answers = tableToField(table, answers)
    return AddAGIFields(answers)

"""
#Say we just want to extract the tables and concatenate them, without regard to what's actually in them or the page Numbers
#The following function is meant to do this and will leave the actual dataframe operations to another wrapper function that you can make

def Table(file, answers):
    file_path = str(file)
    if not Path(file_path).exists():
        raise FileNotFoundError(f"File does not exist: {file_path}")

    pdf = PDF(file_path, detect_rotation=False, pdf_text_extraction=True)
    ocr = TesseractOCR(lang="eng")
    tables = pdf.extract_tables(ocr=ocr)

    pages = list(tables.keys()) if tables else []
    
    if not pages:
        print("Warning: No tables found in PDF")
        return answers
    pdf.to_xlsx('tables.xlsx', ocr=ocr)
    
    table = TableJoinerIndiscriminate(tables, pages)
    table.to_excel("output_table.xlsx", index=False)
    answers = TableToFields(table, answers) #This is for someone to implement based on the specific pdf they have
    return answers

def TableJoinerIndiscriminate(tables, pages):
    FullDF = pd.DataFrame() #We build onto this
    
    for page_num in pages: #This loop 'vertically' joins the tables for each page
        if page_num not in tables:
            continue
        page_df = pd.DataFrame()
        page_tables = tables[page_num]
        
        for table in page_tables: #This loop 'horizontally' joins the tables on the current page
            frame = table.df.copy()
            if frame.empty: #skip empty tables
                continue
            frame.columns = frame.columns.astype(str) #convert columns to string and clean
            frame = frame.dropna(how='all') #Remove rows that are entirely NaN or empty
            if frame.empty: #If frame continued nothing, skip
                continue
            if page_df.empty:
                page_df = frame #Horizontally concatenate tables within the same page
            else:
                page_df = pd.concat([page_df, frame], axis=1, ignore_index=False, sort=False) # Use outer join to preserve all columns
        
        #Remove duplicate columns that might have been created during horizontal concatenation
        if not page_df.empty:
            page_df = page_df.loc[:, ~page_df.columns.duplicated()]
        #Vertically concatenate pages
        if not page_df.empty:
            FullDF = pd.concat([FullDF, page_df], axis=0, ignore_index=True, sort=False)
    FullDF = FullDF.dropna(axis=1, how='all') #remove completely empty columns
    return FullDF
    
"""

#If we have multiple orders in the Unit column, we need to include that so we can cross-reference 
def tableToFields(df, answers, columnToReference="unit"):
    # VERY IMPORTANT!!! column names to lowercase
    df.columns = df.columns.str.lower()
    
    # Convert possibleColumns keys to lowercase for matching
    possibleColumns_lower = {k.lower(): v for k, v in possibleColumns.items()}
    
    match = re.search(r'(\d+\s[\w\s]+),', answers["Address"]) #Get addresses from regex
    if match:
        cleaned_address = match.group(1).strip() #We should always get an address
    else: #If we don't get an address for whatever reason (perhaps the regex wasn't good enough)
        cleaned_address = "" #we give it the empty string to effectively ignore Unit
    print(cleaned_address)
    
    if columnToReference in df.columns and cleaned_address: #For each non-empty address
        df_filtered = df[df[columnToReference].astype(str).str.contains(re.escape(cleaned_address), case=False, na=False)]
        # If no matches found after filtering, use the full dataframe
        if df_filtered.empty:
            print(f"Warning: No matches found for address '{cleaned_address}' in unit column. Using full table.")
            df_filtered = df
    else:
        df_filtered = df
    df_filtered.to_excel("output_table.xlsx", index=False) #You may inspect the table at this file address
    
    for column in df_filtered.columns: #Look at the below function
        if column in possibleColumns_lower and possibleColumns_lower[column] not in answers.keys():
            numeric_series = pd.to_numeric(df_filtered[column], errors='coerce').dropna()
            if not numeric_series.empty:
                answers[possibleColumns_lower[column]] = numeric_series.value_counts().to_dict()
    return answers

#If just a single address, we ignore Unit entirely and simply summarize the table
def tableToField(df, answers):
    for column in df.columns: #For each column
        if column in possibleColumns and possibleColumns[column] not in answers.keys(): #if it's in what we want to add and not already part of answers
            numeric_series = pd.to_numeric(df[column], errors='coerce').dropna() #Remove fields that could cause errors
            answers[possibleColumns[column]] = numeric_series.value_counts().to_dict() #Count up the instances of each rate/useful life
    return answers

def tableJoiner(tables, pages):
    FullDF = pd.DataFrame()
    for i in pages:
        df = pd.DataFrame()
        
        year = {"TotalIncreaseexcludesguideline":0,
                    "%incforCapExp":0}
        for j in tables[i]:
            removed = 0
            frame = j.df
            frame.columns = frame.columns.astype(str).str.replace(r'[()\d\s.]+', '', regex=True)
            while (not any(col in frame.columns for col in splitColumns + totalColumns) and removed < 5 and len(frame) > 0):
                frame.columns = frame.iloc[0]
                frame = frame[1:]
                frame.columns = frame.columns.astype(str).str.replace(r'[()\d\s.]+', '', regex=True)
                removed += 1

            #this regex simply removes digits, whitespace, and .'s.
            frame.columns = frame.columns.astype(str).str.replace(r'[()\d\s.]+', '', regex=True)

            selected_cols = []
            column_map = {}

            for col in frame.columns:
                col_lower = col.lower()
                if col in totalColumns:
                    selected_cols.append(col)

                if col in splitColumns:
                    year[col] = year[col] + 1
                    column_map[col] = f"{col}{year[col]}"
                    selected_cols.append(col)

                if "unit" in col_lower:
                    selected_cols.append(col)
                    
            frame = frame[selected_cols]
            frame = frame.rename(columns = column_map)

            for col in frame.columns:
                try:
                    frame[col] = pd.to_numeric(frame[col])
                except Exception:
                    pass
            
            df = pd.concat([df, frame], axis=1)
            df = df.loc[:, ~df.columns.duplicated()]
        FullDF = pd.concat([FullDF, df], axis=0, ignore_index=True)
    return FullDF

#You'll notice that the below function takes a page list and a Schedule Number
#Perhaps you just care about getting the text from a pdf that's probably all text
def getOCR(pages, scheduleNo, file):
    doc = fitz.open(file) #Open the file with a PDF viewer that can pull objects from it
    #Objects in this case are text, images, tables, etc.
    #This can save us from doing heavy processing if it's not scanned
    all_text = "" #Accumulator string

    for page_index in pages[scheduleNo]: #loop over each page and add to accumulator string
        page = doc[page_index]
        selectable_text = page.get_text().strip() #get_text simply takes all native text that it can find
        #gets text from text

        ocr_text = "" #Now, we perform OCR on 
        for img in page.get_images(full=True): #there can be multiple images (the entire page can be an image!), so we use OCR on each
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image = Im.open(io.BytesIO(image_bytes))
            ocr_text_part = pytesseract.image_to_string(image).strip() #gets text from image
            if ocr_text_part: #if there is text, append
                ocr_text += ocr_text_part + "\n"
            image.close() #otherwise we get overhead and risk file errors

        combined = selectable_text 
        if ocr_text: #add any image text to the text text, if image text is present
            if ocr_text not in selectable_text:
                combined = (selectable_text + "\n" + ocr_text).strip()
        all_text = all_text + combined + "\n"
    doc.close()
    
    if scheduleNo != 3: #I don't want to print tables - use Table functions for that
        print(all_text)
    return all_text

""" #You can use this if you're just passing the entire thing to a SLM
def getOCR(file):
    doc = fitz.open(file)
    all_text = ""
    
    for page_index in range(len(doc)):  # Loop over every page in the document
        page = doc[page_index]
        selectable_text = page.get_text().strip()
        # gets text from text
        ocr_text = ""
        
        for img in page.get_images(full=True):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image = Im.open(io.BytesIO(image_bytes))
            ocr_text_part = pytesseract.image_to_string(image).strip()
            if ocr_text_part:  # if there is text, append
                ocr_text += ocr_text_part + "\n"
            image.close()
            
        combined = selectable_text 
        if ocr_text:
            if ocr_text not in selectable_text:
                combined = (selectable_text + "\n" + ocr_text).strip()
        all_text = all_text + combined + "\n"
        
    doc.close()
    print(all_text)
    return all_text
"""

#Regex module - we try to be opportunistic and hope the answers are within a pattern of some sort
def Regex(text, answers):
    for regex in regexDict.keys(): #for each regular expression
        answer = regexSmall(text, regex) #
        if answer: #if we have an answer, add it
            answers[regexDict[regex]] = answer

        #if we have all the answers, exit loop
        if set(regexDict.values()).issubset(answers.keys()):
            break
            
        #check if dictionary contains all fields, break if so
        #in the worst case when the regexDict ever gets large (unlikely)
        #significant room for optimization if regexDict ever does get large
        #fortunately, this is extremely unlikely for now
    return answers

def regexSmall(text, pattern):
    match = re.search(pattern, text, re.DOTALL) #search for the regex in the text
    if match:
        if len(match.groups()) > 1: #if we have multiple groups that match, concatenate
            result = ' '.join(match.groups())
        else:
            result = match.group(1) #else, use the only match
        result = re.sub(r'\s+', ' ', result.replace('\n', ' '))
        return result.strip()
    return None #if no match, return nothing

#addRow appends a record (JSON) to an excel file (databasePath), under sheetName ('sample') 
def addRow(JSON, databasePath, sheetName, allowRepeats = True):
    wb = openpyxl.load_workbook(databasePath)

    #handle cases for sheet potentially not existing
    if sheetName not in wb.sheetnames or not allowRepeats: #if sheet doesn't exist or we don't allow repeats
        ws = wb.create_sheet(title=sheetName)
        ws.append(list(totalFields))
    else:
        ws = wb[sheetName]

    if ws.max_row == 0 or all(cell.value is None for cell in ws[1]): #if empty sheet, initialize the sheet with our column names
        ws.append(list(totalFields)) #column names are important so we can read it out later!
        #I recommend adding your fields manually though to preserve order, as this can swap them around and make it annoying for humans to analyze

    for col in totalFields: #for each answer missing from our answer dictionary,
        if col not in JSON:
            JSON[col] = "" #give it an empty string to avoid errors when writing (some null types throw errors)

    newRow = []
    for col in totalFields:
        val = JSON[col]
        if isinstance(val, dict): #you may want to later replace this with something such as
        """if !isinstance(val, str):
            val = str(val)""" #this will turn EVERYTHING into a string
            val = str(val) #turn dictionaries into a string so openpyxl doesn't throw type errors
        newRow.append(val)

    ws.append(newRow) #add to excel
    wb.save(databasePath) #save excel file (otherwise it looks like no change was recorded)
    print(f"Appended {JSON.get('FileNumber', '[no FileNumber]')} to {databasePath}")

"""
This function checks if we've already added the file to Excel - we probably don't want to operate on a file if we already have it.
During testing if you don't want to reinitialize the excel files, just replace with "return False"
"""
def checkExcel(dbPath, file_name):
    df = pd.read_excel(dbPath) #Store the excel file as a dataframe
    if df.empty or 'FileName' not in df.columns: #If it's an empty file or doesn't have a FileName column, assume it's not in there
        return False
    match_df = df[df['FileName'] == file_name] #We make a dataframe consisting of one row where FileName is what we're looking for
    return not match_df.empty #if that row exists in this dataframe, we know we have it in the excel file, and vice versa 

def displayDict(dct): #Just for testing, debugging, and viewing the answer file at specific points
    for k,v in dct.items():
        print(f"{k}: {v}")
    print("---------------------------------------------------------------")

"""
This function wraps our main function.
We take a file object, two folders (folderPath = where we get pdfs from), (newFolderPath = where we put pdfs when finished)
another folder (errorFolderPath = where we put hard-to-extract pdfs/pdfs that we encounter an error with)
and two excel files (the new and local DatabasePath, where we put our reocrds)
"""
def FinalApproachWrapper(file, folderPath, newFolderPath, localDatabasePath, newDatabasePath, errorFolderPath):
    sheetName = "sample"
    try: #try to use finalApproach
        print(file.stem)
        if not checkExcel(localDatabasePath, file.stem): #check that the file isn't already in the excel
            start = time.perf_counter() #these time. functions are so I can measure performance - you may remove them
            answers = FinalApproach(file) #get answers
            for answer in answers:
                displayDict(answer) #display our answers
            for x in answers:
                if "toAdd" not in x.keys(): #meant for someone to add a field that would indicate an invalid answer dictionary, but we currently do nothing with this
                    n = addRow(x, localDatabasePath, sheetName) #add 'JSON' to excel files
                    m = addRow(x, newDatabasePath, sheetName) #we currently add to both, you can play around with this if you want to change where we're moving files

            end = time.perf_counter()
            elapsed = end - start
            print(f"Total execution time: {elapsed:.2f} seconds")
        shutil.move(str(file), newFolderPath / file.name) #no errors? put the file into the finished folder
    except Exception as e: #if we get an error anywhere, throw the pdf into the error folder
        print(f"Error processing {file.name}: {e}")
        shutil.move(str(file), errorFolderPath / file.name)
        

def main():
    print("Entering main")
    sheetName = 'sample'
    if True: #this branch prompts the user to provide the file paths without having to 
        root = tk.Tk()
        root.withdraw()
        root.update()
        
        folderPath = Path(filedialog.askdirectory()) #ask for folder to get pdfs from
        f = filedialog.askopenfilenames() #ask for two excel files to add to
        localDatabasePath, newDatabasePath = Path(f[0]), Path(f[1])
        newFolderPath = Path(filedialog.askdirectory()) #ask for folder to put pdfs into once finished
    """
    #You can set these as such if you're repeatedly running this script
    newFolderPath = Path(r"C:\Users\marut\OneDrive\Desktop\..UofT\Second Year\RentAGI\AGI Database Work Summer 2025\ReadOrders")
    """

    parentDir = newFolderPath.parent 
    errorFolderPath = parentDir / "Excessively long" #folder called Excessively long, for files we encounter errors with
    errorFolderPath.mkdir(parents=True, exist_ok=True) #designate and make folder next to the folder that we're adding to

    start_time = time.time()
    time_array = [] #for tracking
    print("Process started.")

    folder = [file for file in folderPath.glob("*.pdf") if (("Order" in file.name) or ("Hearing" in file.name) or ("Redacted" in file.name))]
    """lazy match for pdfs that we want to work on
    I think it's worth adding a couple conditions (perhaps with functions) so you don't add everything, but assuming this is a folder
    with just court orders, you don't need the conditions.
    If you want to just work on every pdf (or other file types)
    folder = [file for file in folderPath.glob("*.pdf")], 
    """

    for file in folder[:]: #for every file in the folder, run final approach.
        #folder[:] is just a 'list' of every file, so you can pick and choose/manipulate the order of pdfs that we extract data from
        FinalApproachWrapper(file, folderPath, newFolderPath, localDatabasePath, newDatabasePath, errorFolderPath)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(time_array)
    print(f"Total execution time: {elapsed_time:.2f} seconds")    
    
if __name__ == "__main__": #runs the following code. You can wrap main here or declare additional constants
    singleModel = models[0] #set SLM here
    main()
